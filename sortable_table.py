"""
네이버 카페 회원 관리 - 정렬 가능한 테이블
- 다중 컬럼 정렬 지원 (A열 클릭 → B열 클릭 → A 유지, B 세부 정렬)
- 클릭 한 번: 정순 ↓ / 두 번: 역순 ↑ / 세 번: 해제 ↕
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import csv
import json
import os
import re
import locale
import datetime
import calendar
import threading

# 한국어 정렬을 위한 locale 설정
try:
    locale.setlocale(locale.LC_COLLATE, 'ko_KR.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_COLLATE, 'Korean_Korea.949')
    except locale.Error:
        pass

NEUTRAL = '\u2195'  # ↕
ASC = '\u2193'      # ↓
DESC = '\u2191'     # ↑

# 정렬 타입
SORT_TYPE_TEXT = 'text'
SORT_TYPE_NUMBER = 'number'
SORT_TYPE_BOOL = 'bool'
SORT_TYPE_COMPOSITE = 'composite'  # "3(1/2)" 형식 → 앞의 합계로 정렬

# 고정 컬럼 정의 (9열)
FIXED_COLUMNS = [
    '네이버ID', '카페닉네임', '본명', '동행천만', '밥상모임',
    '게시글수', '출석체크', '수강생 후기(주차별/라스트원)', '보충강의', '라이브 후기'
]
FIXED_SORT_TYPES = {
    '네이버ID': SORT_TYPE_TEXT,
    '카페닉네임': SORT_TYPE_TEXT,
    '본명': SORT_TYPE_TEXT,
    '동행천만': SORT_TYPE_BOOL,
    '밥상모임': SORT_TYPE_BOOL,
    '게시글수': SORT_TYPE_NUMBER,
    '출석체크': SORT_TYPE_NUMBER,
    '수강생 후기(주차별/라스트원)': SORT_TYPE_COMPOSITE,
    '보충강의': SORT_TYPE_NUMBER,
    '라이브 후기': SORT_TYPE_BOOL,
}
COL_IDX = {name: i for i, name in enumerate(FIXED_COLUMNS)}


def sort_key_for_type(value, sort_type):
    """타입별 정렬 키 생성"""
    v_str = str(value).strip()

    if sort_type == SORT_TYPE_NUMBER:
        try:
            return (0, float(v_str.replace(',', '')))
        except (ValueError, TypeError):
            return (1, v_str)

    if sort_type == SORT_TYPE_BOOL:
        negative = {'', 'x', '없음', 'no', 'false'}
        if v_str.lower() in negative:
            return (0, 1)
        return (0, 0)

    if sort_type == SORT_TYPE_COMPOSITE:
        # "3(1/2)" → 앞의 숫자(합계)로 정렬
        match = re.match(r'^(\d+)', v_str)
        if match:
            return (0, float(match.group(1)))
        return (1, v_str)

    return (0, v_str)


class SortableTable:
    def __init__(self, root):
        self.root = root
        self.root.title("카페 회원 관리 - 정렬 테이블")
        self.root.geometry("1450x650")
        self.root.configure(bg='#1a1a2e')

        self.columns = list(FIXED_COLUMNS)
        self.sort_types = dict(FIXED_SORT_TYPES)
        self.data = []
        self.original_data = []
        self.sort_order = []  # [(column, direction), ...] 다중 정렬
        self.row_flags = {}
        # 필터 모드: 열별 '=' (전체), '∩' (있음-교집합), '∪' (있음-합집합),
        #           '!∩' (없음-교집합), '!∪' (없음-합집합)
        self.filter_modes = {}  # {column: mode_str}
        # 네이버ID, 카페닉네임은 필터 대상 아님
        self.filterable_cols = [c for c in FIXED_COLUMNS if c not in ('네이버ID', '카페닉네임')]

        self.data_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'member_data.json')
        self.col_config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'column_config.json')
        self.col_config = self._load_col_config()

        self._build_ui()
        self._setup_fixed_columns()
        self._load_persisted_data()

    def _build_ui(self):
        """UI 전체 구성 (툴바, 테이블, 필터, 상태바) / Build the entire UI layout"""
        # === 툴바 1줄: 메인 기능 ===
        toolbar_top = tk.Frame(self.root, bg='#16213e', pady=6, padx=10)
        toolbar_top.pack(fill='x')

        tk.Label(toolbar_top, text="카페 회원 관리", font=('맑은 고딕', 14, 'bold'),
                 bg='#16213e', fg='#e94560').pack(side='left', padx=(0, 15))

        btn_main = {'font': ('맑은 고딕', 9), 'bg': '#0f3460', 'fg': 'white',
                    'relief': 'flat', 'padx': 8, 'pady': 3, 'cursor': 'hand2'}
        btn_scrape = {'font': ('맑은 고딕', 9, 'bold'), 'bg': '#e94560', 'fg': 'white',
                      'relief': 'flat', 'padx': 8, 'pady': 3, 'cursor': 'hand2'}
        btn_danger = {'font': ('맑은 고딕', 9), 'bg': '#8b0000', 'fg': 'white',
                      'relief': 'flat', 'padx': 8, 'pady': 3, 'cursor': 'hand2'}

        tk.Button(toolbar_top, text="회원 조회",
                  command=self._scrape_members_only, **btn_scrape).pack(side='left', padx=2)
        tk.Button(toolbar_top, text="게시판 집계",
                  command=self._scrape_boards_only, **btn_scrape).pack(side='left', padx=2)
        btn_config = {'font': ('맑은 고딕', 9), 'bg': '#2d6a4f', 'fg': 'white',
                      'relief': 'flat', 'padx': 8, 'pady': 3, 'cursor': 'hand2'}
        tk.Button(toolbar_top, text="게시판 변수 설정",
                  command=self._open_board_config, **btn_config).pack(side='left', padx=2)
        tk.Button(toolbar_top, text="카페 JSON 불러오기",
                  command=self._load_cafe_json, **btn_main).pack(side='left', padx=2)
        tk.Button(toolbar_top, text="CSV로 저장",
                  command=self._save_csv, **btn_main).pack(side='left', padx=2)
        tk.Button(toolbar_top, text="행 추가",
                  command=self._add_row, **btn_main).pack(side='left', padx=2)
        tk.Button(toolbar_top, text="행 삭제",
                  command=self._delete_row, **btn_main).pack(side='left', padx=2)
        tk.Button(toolbar_top, text="모든 행 삭제",
                  command=self._delete_all_rows, **btn_danger).pack(side='left', padx=2)
        tk.Button(toolbar_top, text="정렬 초기화",
                  command=self._reset_sort, **btn_main).pack(side='left', padx=2)

        self.sort_label = tk.Label(toolbar_top, text="정렬: 없음", font=('맑은 고딕', 9),
                                   bg='#16213e', fg='#a8a8a8')
        self.sort_label.pack(side='right')

        # === 툴바 2줄: CSV 속성 매핑 ===
        toolbar_bottom = tk.Frame(self.root, bg='#1a1a3e', pady=5, padx=10)
        toolbar_bottom.pack(fill='x')

        tk.Label(toolbar_bottom, text="속성 매핑:", font=('맑은 고딕', 9),
                 bg='#1a1a3e', fg='#a8a8a8').pack(side='left', padx=(0, 8))

        btn_csv = {'font': ('맑은 고딕', 9), 'bg': '#1a5276', 'fg': 'white',
                   'relief': 'flat', 'padx': 8, 'pady': 3, 'cursor': 'hand2'}
        btn_util = {'font': ('맑은 고딕', 9), 'bg': '#4a3f00', 'fg': 'white',
                    'relief': 'flat', 'padx': 8, 'pady': 3, 'cursor': 'hand2'}

        tk.Button(toolbar_bottom, text="동행천만 CSV",
                  command=self._load_dongheng_csv, **btn_csv).pack(side='left', padx=2)
        tk.Button(toolbar_bottom, text="밥상모임 CSV",
                  command=self._load_babsang_csv, **btn_csv).pack(side='left', padx=2)
        tk.Button(toolbar_bottom, text="본명 CSV",
                  command=self._load_realname_csv, **btn_csv).pack(side='left', padx=2)
        tk.Frame(toolbar_bottom, bg='#1a1a3e', width=20).pack(side='left')
        btn_col_cfg = {'font': ('맑은 고딕', 9), 'bg': '#2d6a4f', 'fg': 'white',
                       'relief': 'flat', 'padx': 8, 'pady': 3, 'cursor': 'hand2'}
        tk.Button(toolbar_bottom, text="열 이름 설정",
                  command=self._open_col_config, **btn_col_cfg).pack(side='left', padx=2)
        tk.Button(toolbar_bottom, text="하이라이트 초기화",
                  command=self._clear_highlights, **btn_util).pack(side='left', padx=2)

        # === 테이블 영역 ===
        table_frame = tk.Frame(self.root, bg='#1a1a2e')
        table_frame.pack(fill='both', expand=True, padx=10, pady=(5, 10))

        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Treeview', background='#16213e', foreground='white',
                        fieldbackground='#16213e', font=('맑은 고딕', 10), rowheight=28)
        style.configure('Treeview.Heading', background='#0f3460', foreground='white',
                        font=('맑은 고딕', 10, 'bold'), relief='flat')
        style.map('Treeview', background=[('selected', '#e94560')],
                  foreground=[('selected', 'white')])
        style.map('Treeview.Heading', background=[('active', '#e94560')])

        self.tree = ttk.Treeview(table_frame, show='headings', selectmode='extended')
        scrollbar_y = ttk.Scrollbar(table_frame, orient='vertical', command=self.tree.yview)
        scrollbar_x = ttk.Scrollbar(table_frame, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        self.tree.grid(row=0, column=0, sticky='nsew')
        scrollbar_y.grid(row=0, column=1, sticky='ns')
        scrollbar_x.grid(row=1, column=0, sticky='ew')
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # 더블클릭으로 셀 편집
        self.tree.bind('<Double-1>', self._on_cell_double_click)
        self._edit_widget = None
        # 헤더 클릭: 정렬
        self.tree.bind('<Button-1>', self._on_tree_click)
        # 우클릭: 열 이동 메뉴
        self.tree.bind('<Button-3>', self._on_header_right_click)

        # === 필터 버튼 행 ===
        filter_frame = tk.Frame(self.root, bg='#0d1b2a', pady=3, padx=10)
        filter_frame.pack(fill='x')
        tk.Label(filter_frame, text="필터:", font=('맑은 고딕', 9),
                 bg='#0d1b2a', fg='#a8a8a8').pack(side='left', padx=(0, 5))
        self.filter_buttons = {}
        for col in self.columns:
            if col not in self.filterable_cols:
                continue
            # 짧은 열 이름
            short = col[:4] if len(col) > 6 else col
            btn = tk.Button(filter_frame, text=f"{short}=",
                            font=('맑은 고딕', 8), bg='#1b2838', fg='#a8a8a8',
                            relief='flat', padx=4, pady=1, cursor='hand2',
                            command=lambda c=col: self._cycle_filter(c))
            btn.pack(side='left', padx=1)
            self.filter_buttons[col] = btn
        # 필터 초기화 버튼
        tk.Button(filter_frame, text="초기화", font=('맑은 고딕', 8),
                  bg='#8b0000', fg='white', relief='flat', padx=4, pady=1,
                  cursor='hand2', command=self._reset_filters).pack(side='left', padx=(8, 0))

        # === 하단 상태바 ===
        status_bar = tk.Frame(self.root, bg='#0f3460', pady=4)
        status_bar.pack(fill='x')
        self.status_label = tk.Label(status_bar, text="준비됨", font=('맑은 고딕', 9),
                                     bg='#0f3460', fg='#a8a8a8')
        self.status_label.pack(side='left', padx=10)
        self.count_label = tk.Label(status_bar, text="0명", font=('맑은 고딕', 9),
                                    bg='#0f3460', fg='#a8a8a8')
        self.count_label.pack(side='right', padx=10)

    # ──────────────────────────────────────────────
    # 컬럼 / 데이터 설정
    # ──────────────────────────────────────────────

    def _setup_fixed_columns(self):
        """고정 컬럼 헤더 및 너비 초기 설정 / Initialize fixed column headings and widths"""
        self.tree.configure(columns=self.columns)
        for col in self.columns:
            self.tree.heading(col, text=f"{col} {NEUTRAL}")
            if col in ('네이버ID', '카페닉네임'):
                width = 130
            elif col == '수강생 후기(주차별/라스트원)':
                width = 180
            else:
                width = 90
            self.tree.column(col, width=width, minwidth=60, anchor='center')

    def _set_data(self, data):
        """데이터 일괄 설정 및 테이블 갱신 / Set all row data and refresh the table"""
        self.data = [list(row) for row in data]
        self.original_data = [list(row) for row in data]
        self._refresh_table()
        self._update_count()

    # ──────────────────────────────────────────────
    # 다중 컬럼 정렬 + 필터 모드
    # ──────────────────────────────────────────────

    FILTER_CYCLE = ['=', '∩', '∪', '!∩', '!∪']

    def _on_tree_click(self, event):
        """좌클릭: 정렬 / Ctrl+좌클릭: 필터 모드 전환"""
        region = self.tree.identify_region(event.x, event.y)
        if region != 'heading':
            return  # 헤더가 아니면 기본 동작 (행 선택 등)

        column_id = self.tree.identify_column(event.x)
        if not column_id:
            return
        col_idx = int(column_id.replace('#', '')) - 1
        if col_idx < 0 or col_idx >= len(self.columns):
            return
        col_name = self.columns[col_idx]

        # Ctrl 키 비트마스크: 0x4 (Ctrl이 눌린 상태인지 확인)
        if event.state & 0x4:
            # Ctrl+클릭: 필터 모드 순환 전환
            if col_name in self.filterable_cols:
                current = self.filter_modes.get(col_name, '=')
                idx = self.FILTER_CYCLE.index(current) if current in self.FILTER_CYCLE else 0
                next_mode = self.FILTER_CYCLE[(idx + 1) % len(self.FILTER_CYCLE)]
                self._set_filter_mode(col_name, next_mode)
        else:
            # 일반 클릭: 정렬
            self._on_header_click(col_name)

    def _on_header_click(self, column):
        """헤더 클릭: ASC->DESC->제거, 새 열은 sort_order에 추가"""
        # 이미 정렬 중인 열인지 확인
        existing_idx = None
        existing_dir = None
        for i, (col, direction) in enumerate(self.sort_order):
            if col == column:
                existing_idx = i
                existing_dir = direction
                break

        if existing_idx is not None:
            # 정순이면 역순으로 전환, 역순이면 정렬에서 제거 (3단계 순환)
            if existing_dir == ASC:
                self.sort_order[existing_idx] = (column, DESC)
            else:
                self.sort_order.pop(existing_idx)
        else:
            # 새로운 열: 정순으로 추가 (다중 정렬)
            self.sort_order.append((column, ASC))

        self._update_header_labels()
        self._apply_sort()

    def _update_header_labels(self):
        """모든 컬럼 헤더에 정렬 화살표와 필터 모드 표시 갱신 / Refresh heading arrows and filter indicators"""
        for col in self.columns:
            # 정렬 화살표 결정
            arrow = NEUTRAL
            for sort_col, direction in self.sort_order:
                if sort_col == col:
                    arrow = ASC if direction == ASC else DESC
                    break
            # 필터 모드
            fmode = self.filter_modes.get(col, '=')
            filter_txt = '' if fmode == '=' or col not in self.filterable_cols else fmode
            self.tree.heading(col, text=f"{col} {arrow}{filter_txt}")

    def _apply_sort(self):
        """현재 sort_order 기준으로 다중 컬럼 정렬 실행 / Apply multi-column sort based on sort_order"""
        if not self.sort_order:
            # 정렬 조건이 없으면 원본 순서로 복원
            self.data = [list(row) for row in self.original_data]
            self.sort_label.configure(text="정렬: 없음")
        else:
            self.data = [list(row) for row in self.original_data]
            # stable sort: 역순으로 적용하여 우선순위 보장 (마지막 키부터 정렬)
            for col, direction in reversed(self.sort_order):
                col_idx = COL_IDX[col]
                s_type = self.sort_types[col]
                reverse = (direction == DESC)
                self.data.sort(
                    key=lambda row, ci=col_idx, st=s_type: sort_key_for_type(
                        row[ci] if ci < len(row) else '', st
                    ),
                    reverse=reverse
                )
            # 상태 라벨에 현재 정렬 순서 표시 (예: "정렬: 카페닉네임↓, 게시글수↑")
            parts = []
            for col, direction in self.sort_order:
                arrow = "↓" if direction == ASC else "↑"
                parts.append(f"{col}{arrow}")
            self.sort_label.configure(text=f"정렬: {', '.join(parts)}")
        self._refresh_table()

    def _reset_sort(self):
        """모든 정렬 조건 초기화 및 원본 순서 복원 / Clear all sort orders and restore original row order"""
        self.sort_order = []
        self._update_header_labels()
        self.data = [list(row) for row in self.original_data]
        self.sort_label.configure(text="정렬: 없음")
        self._refresh_table()

    # ──────────────────────────────────────────────
    # 테이블 표시
    # ──────────────────────────────────────────────

    def _has_value(self, cell_val):
        """셀에 값이 있는지 (빈 문자열, '0', '0(0/0)', 'X' = 없음)"""
        v = str(cell_val).strip()
        return v not in ('', '0', '0(0/0)', 'X')

    def _passes_filter(self, row):
        """현재 필터 모드에 따라 이 행이 표시되어야 하는지 판정"""
        if not self.filter_modes:
            return True

        # 필터가 적용된 열들을 모드별로 분류
        union_cols = []     # ∪: 하나라도 있으면 통과
        intersect_cols = [] # ∩: 모두 있어야 통과
        not_union_cols = [] # !∪: 하나라도 없으면 통과
        not_intersect_cols = []  # !∩: 모두 없어야 통과

        for col, mode in self.filter_modes.items():
            if mode == '=' or col not in COL_IDX:
                continue
            if mode == '∪':
                union_cols.append(col)
            elif mode == '∩':
                intersect_cols.append(col)
            elif mode == '!∪':
                not_union_cols.append(col)
            elif mode == '!∩':
                not_intersect_cols.append(col)

        # 각 조건 평가
        results = []

        if intersect_cols:
            # ∩: 모든 열에 값이 있어야 함
            results.append(all(self._has_value(row[COL_IDX[c]]) for c in intersect_cols if COL_IDX[c] < len(row)))

        if union_cols:
            # ∪: 하나라도 값이 있으면 됨
            results.append(any(self._has_value(row[COL_IDX[c]]) for c in union_cols if COL_IDX[c] < len(row)))

        if not_intersect_cols:
            # !∩: 모든 열에 값이 없어야 함
            results.append(all(not self._has_value(row[COL_IDX[c]]) for c in not_intersect_cols if COL_IDX[c] < len(row)))

        if not_union_cols:
            # !∪: 하나라도 값이 없으면 됨
            results.append(any(not self._has_value(row[COL_IDX[c]]) for c in not_union_cols if COL_IDX[c] < len(row)))

        return all(results) if results else True

    def _refresh_table(self):
        """Treeview 전체 행 다시 그리기 (필터 적용 포함) / Redraw all Treeview rows with filters applied"""
        # 기존 행 모두 삭제
        for item in self.tree.get_children():
            self.tree.delete(item)
        visible_count = 0
        for i, row in enumerate(self.data):
            # 필터 조건에 맞지 않는 행은 건너뛰기
            if not self._passes_filter(row):
                continue
            visible_count += 1
            nid = row[COL_IDX['네이버ID']] if len(row) > 0 else ''
            flags = self.row_flags.get(nid, {})
            # 행 상태에 따라 태그 결정 (신규/갱신/짝수줄/홀수줄)
            if flags.get('new'):
                tag = 'new_row'
            elif flags.get('updated'):
                tag = 'updated_row'
            elif visible_count % 2 == 1:
                tag = 'even'
            else:
                tag = 'odd'
            self.tree.insert('', 'end', values=row, tags=(tag,))
        # 태그별 배경색 설정
        self.tree.tag_configure('even', background='#16213e')
        self.tree.tag_configure('odd', background='#1a1a3e')
        self.tree.tag_configure('new_row', background='#4a3f00')
        self.tree.tag_configure('updated_row', background='#1a3a2e')

    def _update_count(self):
        """하단 상태바의 회원 수 표시 갱신 / Update the member count label in the status bar"""
        total = len(self.data)
        # 필터가 활성화된 경우 보이는 행 수 / 전체 수 형태로 표시
        visible = sum(1 for row in self.data if self._passes_filter(row))
        if visible < total and self.filter_modes:
            self.count_label.configure(text=f"{visible}/{total}명")
        else:
            self.count_label.configure(text=f"{total}명")

    def _clear_highlights(self):
        self.row_flags.clear()
        self._refresh_table()
        self.status_label.configure(text="하이라이트 초기화됨")

    # ──────────────────────────────────────────────
    # 우클릭 메뉴 (필터 모드 + 열 이동)
    # ──────────────────────────────────────────────

    def _on_header_right_click(self, event):
        """헤더 우클릭 → 필터 모드 선택 / 열 이동 메뉴"""
        region = self.tree.identify_region(event.x, event.y)
        if region != 'heading':
            return
        column = self.tree.identify_column(event.x)
        if not column:
            return
        col_idx = int(column.replace('#', '')) - 1
        if col_idx < 0 or col_idx >= len(self.columns):
            return
        col_name = self.columns[col_idx]

        menu = tk.Menu(self.root, tearoff=0, bg='#16213e', fg='white',
                       activebackground='#e94560', activeforeground='white',
                       font=('맑은 고딕', 9))

        # 필터 모드 (네이버ID, 카페닉네임 제외)
        if col_name in self.filterable_cols:
            current_mode = self.filter_modes.get(col_name, '=')
            menu.add_command(label=f"── 필터: {col_name} ──", state='disabled')
            for mode, desc in [('=', '= 전체 표시'),
                               ('∩', '∩ 있음 (교집합)'),
                               ('∪', '∪ 있음 (합집합)'),
                               ('!∩', '!∩ 없음 (교집합)'),
                               ('!∪', '!∪ 없음 (합집합)')]:
                label = f"{'● ' if current_mode == mode else '  '}{desc}"
                menu.add_command(label=label,
                                 command=lambda m=mode, c=col_name: self._set_filter_mode(c, m))
            menu.add_separator()

        # 열 이동
        menu.add_command(label=f"── 열 이동: {col_name} ──", state='disabled')
        if col_idx > 0:
            menu.add_command(label="◀ 왼쪽으로 이동",
                             command=lambda: self._move_column(col_idx, col_idx - 1))
        if col_idx < len(self.columns) - 1:
            menu.add_command(label="▶ 오른쪽으로 이동",
                             command=lambda: self._move_column(col_idx, col_idx + 1))

        menu.add_separator()
        menu.add_command(label="모든 필터 초기화", command=self._reset_filters)

        menu.post(event.x_root, event.y_root)

    def _cycle_filter(self, col_name):
        """필터 버튼 클릭 → 다음 모드로 전환"""
        current = self.filter_modes.get(col_name, '=')
        idx = self.FILTER_CYCLE.index(current) if current in self.FILTER_CYCLE else 0
        next_mode = self.FILTER_CYCLE[(idx + 1) % len(self.FILTER_CYCLE)]
        self._set_filter_mode(col_name, next_mode)

    def _set_filter_mode(self, col_name, mode):
        """열의 필터 모드 설정"""
        if mode == '=':
            self.filter_modes.pop(col_name, None)
        else:
            self.filter_modes[col_name] = mode
        self._update_header_labels()
        self._update_filter_buttons()
        self._refresh_table()
        self._update_count()
        active = [f"{c}{m}" for c, m in self.filter_modes.items() if m != '=']
        if active:
            self.status_label.configure(text=f"필터: {', '.join(active)}")
        else:
            self.status_label.configure(text="필터 해제됨")

    def _update_filter_buttons(self):
        """필터 버튼 텍스트 + 색상 업데이트 / Sync filter button labels and colors with current modes"""
        # 모드별 배경색/전경색 매핑
        colors = {
            '=': ('#1b2838', '#a8a8a8'),
            '∩': ('#1a5276', '#ffffff'),
            '∪': ('#2d6a4f', '#ffffff'),
            '!∩': ('#8b0000', '#ffffff'),
            '!∪': ('#6b3a00', '#ffffff'),
        }
        for col, btn in self.filter_buttons.items():
            mode = self.filter_modes.get(col, '=')
            # 열 이름이 길면 앞 4글자만 표시
            short = col[:4] if len(col) > 6 else col
            btn.configure(text=f"{short}{mode}")
            bg, fg = colors.get(mode, ('#1b2838', '#a8a8a8'))
            btn.configure(bg=bg, fg=fg)

    def _reset_filters(self):
        """모든 필터 초기화"""
        self.filter_modes.clear()
        self._update_header_labels()
        self._update_filter_buttons()
        self._refresh_table()
        self._update_count()
        self.status_label.configure(text="모든 필터 초기화됨")

    def _move_column(self, from_idx, to_idx):
        """열 순서 변경"""
        if from_idx == to_idx:
            return
        # columns 순서 변경
        col = self.columns.pop(from_idx)
        self.columns.insert(to_idx, col)
        # data, original_data의 각 행에서도 순서 변경
        for row in self.data:
            while len(row) < len(FIXED_COLUMNS):
                row.append('')
            val = row.pop(from_idx)
            row.insert(to_idx, val)
        for row in self.original_data:
            while len(row) < len(FIXED_COLUMNS):
                row.append('')
            val = row.pop(from_idx)
            row.insert(to_idx, val)
        # COL_IDX 재계산
        global COL_IDX
        COL_IDX = {name: i for i, name in enumerate(self.columns)}
        # Treeview 재구성
        self._setup_fixed_columns()
        self._update_header_labels()
        self._refresh_table()
        self._auto_save()
        self.status_label.configure(text=f"열 이동: {col}")

    # ──────────────────────────────────────────────
    # 셀 편집 (더블클릭)
    # ──────────────────────────────────────────────

    def _on_cell_double_click(self, event):
        """더블클릭한 셀 위에 Entry 위젯을 띄워서 편집"""
        # 기존 편집 위젯 정리
        self._cancel_edit()

        # 클릭한 위치에서 행/열 확인
        item = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        if not item or not column:
            return

        # column은 '#1', '#2' 형식 → 인덱스로 변환
        col_idx = int(column.replace('#', '')) - 1
        if col_idx < 0 or col_idx >= len(self.columns):
            return

        # 게시글수~보충강의 열은 편집 불가 (스크래퍼가 관리)
        readonly_cols = {'게시글수', '출석체크', '수강생 후기(주차별/라스트원)', '보충강의'}
        if self.columns[col_idx] in readonly_cols:
            return

        # 셀 위치/크기 가져오기
        bbox = self.tree.bbox(item, column)
        if not bbox:
            return
        x, y, w, h = bbox

        # 현재 값
        current_values = list(self.tree.item(item, 'values'))
        old_value = current_values[col_idx] if col_idx < len(current_values) else ''

        # Entry 위젯 생성
        entry = tk.Entry(self.tree, font=('맑은 고딕', 10),
                         bg='#2a2a4e', fg='white', insertbackground='white',
                         relief='flat', justify='center')
        entry.insert(0, old_value)
        entry.select_range(0, tk.END)
        entry.place(x=x, y=y, width=w, height=h)
        entry.focus_set()

        # 저장 정보
        self._edit_widget = entry
        self._edit_item = item
        self._edit_col_idx = col_idx

        # Enter로 확인, Escape으로 취소, 포커스 벗어나면 확인
        entry.bind('<Return>', lambda e: self._confirm_edit())
        entry.bind('<Escape>', lambda e: self._cancel_edit())
        entry.bind('<FocusOut>', lambda e: self._confirm_edit())

    def _confirm_edit(self):
        """편집 확정 → 데이터 저장"""
        if not self._edit_widget:
            return
        new_value = self._edit_widget.get()
        item = self._edit_item
        col_idx = self._edit_col_idx

        # Treeview 값 업데이트
        current_values = list(self.tree.item(item, 'values'))
        while len(current_values) <= col_idx:
            current_values.append('')
        old_value = current_values[col_idx]

        if new_value != old_value:
            current_values[col_idx] = new_value
            self.tree.item(item, values=current_values)

            # original_data에서도 해당 행 찾아서 업데이트 (정렬과 무관하게 원본 동기화)
            tree_idx = self.tree.index(item)
            if tree_idx < len(self.data):
                self.data[tree_idx] = list(current_values)
                # 원본 데이터에서 변경 전 값 기준으로 일치하는 행을 찾아 새 값 반영
                old_row = self.data[tree_idx]
                for i, row in enumerate(self.original_data):
                    old_vals = [str(v) for v in row]
                    target = list(current_values)
                    target[col_idx] = old_value  # 변경 전 값으로 비교
                    if old_vals == [str(v) for v in target] or row == target:
                        while len(self.original_data[i]) <= col_idx:
                            self.original_data[i].append('')
                        self.original_data[i][col_idx] = new_value
                        break

                self._auto_save()
                self.status_label.configure(
                    text=f"편집됨: {self.columns[col_idx]} = '{new_value}'")

        self._cancel_edit()

    def _cancel_edit(self):
        """편집 위젯 제거"""
        if self._edit_widget:
            self._edit_widget.destroy()
            self._edit_widget = None

    # ──────────────────────────────────────────────
    # Selenium 카페 회원 스크래핑
    # ──────────────────────────────────────────────

    def _load_board_config(self):
        """board_config.json 로드"""
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'board_config.json')
        default = {
            "cafe_name": "lastone0228", "club_id": "31136700",
            "boards": {"출석체크": "51", "수강생 주차별 후기": "105",
                       "라스트원 강의후기": "22", "졸업생 보충강의 후기": "79"}
        }
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception:
                pass
        return default

    def _save_board_config(self, config):
        """게시판 설정을 board_config.json에 저장 / Save board config to board_config.json"""
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'board_config.json')
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)

    def _open_board_config(self):
        """게시판 변수 설정 다이얼로그"""
        config = self._load_board_config()
        dialog = tk.Toplevel(self.root)
        dialog.title("게시판 변수 설정")
        dialog.configure(bg='#1a1a2e')
        dialog.geometry("450x400")
        dialog.transient(self.root)
        dialog.grab_set()

        entries = {}
        fields = [
            ("카페 이름", "cafe_name", config.get("cafe_name", "")),
            ("클럽 ID", "club_id", config.get("club_id", "")),
            ("출석체크 게시판 번호", "출석체크", config.get("boards", {}).get("출석체크", "")),
            ("수강생 주차별 후기 번호", "수강생 주차별 후기", config.get("boards", {}).get("수강생 주차별 후기", "")),
            ("라스트원 강의후기 번호", "라스트원 강의후기", config.get("boards", {}).get("라스트원 강의후기", "")),
            ("졸업생 보충강의 후기 번호", "졸업생 보충강의 후기", config.get("boards", {}).get("졸업생 보충강의 후기", "")),
            ("졸업생 라이브 후기 번호", "졸업생 라이브 후기", config.get("boards", {}).get("졸업생 라이브 후기", "")),
        ]
        for i, (label, key, val) in enumerate(fields):
            tk.Label(dialog, text=label, font=('맑은 고딕', 10),
                     bg='#1a1a2e', fg='white').grid(row=i, column=0, padx=10, pady=5, sticky='e')
            entry = tk.Entry(dialog, font=('맑은 고딕', 10), bg='#16213e', fg='white',
                             insertbackground='white')
            entry.insert(0, val)
            entry.grid(row=i, column=1, padx=10, pady=5, sticky='ew')
            entries[key] = entry
        dialog.grid_columnconfigure(1, weight=1)

        def save():
            new_config = {
                "cafe_name": entries["cafe_name"].get().strip(),
                "club_id": entries["club_id"].get().strip(),
                "boards": {
                    "출석체크": entries["출석체크"].get().strip(),
                    "수강생 주차별 후기": entries["수강생 주차별 후기"].get().strip(),
                    "라스트원 강의후기": entries["라스트원 강의후기"].get().strip(),
                    "졸업생 보충강의 후기": entries["졸업생 보충강의 후기"].get().strip(),
                    "졸업생 라이브 후기": entries["졸업생 라이브 후기"].get().strip(),
                }
            }
            self._save_board_config(new_config)
            self.status_label.configure(text="게시판 설정 저장됨")
            dialog.destroy()

        tk.Button(dialog, text="저장", command=save,
                  font=('맑은 고딕', 10, 'bold'), bg='#2d6a4f', fg='white',
                  relief='flat', padx=20, pady=4).grid(
            row=len(fields), column=0, columnspan=2, pady=12)

    def _scrape_members_only(self):
        """회원 조회만 실행"""
        try:
            from cafe_scraper import CafeScraper
        except ImportError:
            messagebox.showerror("모듈 없음", "cafe_scraper.py + selenium 필요")
            return

        config = self._load_board_config()
        self.status_label.configure(text="회원 조회 준비 중...")
        self.root.update()
        output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'cafe_members.json')

        # 별도 스레드에서 스크래핑 실행 (UI 블로킹 방지)
        def run():
            def on_status(msg):
                # 메인 스레드에서 상태 라벨 업데이트
                self.root.after(0, lambda: self.status_label.configure(text=msg))
            scraper = CafeScraper(config['cafe_name'], on_status=on_status)
            members = scraper.run_full(output_path, mode='members', board_config=config)
            self.root.after(0, lambda: self._on_scrape_complete(members, output_path))

        threading.Thread(target=run, daemon=True).start()

    def _scrape_boards_only(self):
        """게시판 집계만 실행 (출석체크 + 후기 + 보충강의)"""
        try:
            from cafe_scraper import CafeScraper
        except ImportError:
            messagebox.showerror("모듈 없음", "cafe_scraper.py + selenium 필요")
            return

        config = self._load_board_config()

        start_date_str = simpledialog.askstring(
            "시작일 입력",
            "게시판 집계 시작일을 입력하세요:\n(예: 2026-03-14)",
            initialvalue=datetime.datetime.now().strftime('%Y-%m-%d'),
            parent=self.root)
        if not start_date_str or not start_date_str.strip():
            return
        start_date_str = start_date_str.strip()
        if not re.match(r'^\d{4}-\d{2}-\d{2}$', start_date_str):
            messagebox.showwarning("형식 오류", "시작일 형식은 YYYY-MM-DD이어야 합니다.")
            return

        end_date_str = simpledialog.askstring(
            "종료일 입력",
            f"게시판 집계 종료일을 입력하세요:\n(예: 2026-04-13)\n\n"
            f"시작일: {start_date_str}",
            initialvalue=(datetime.date.fromisoformat(start_date_str)
                          + datetime.timedelta(days=30)).isoformat(),
            parent=self.root)
        if not end_date_str or not end_date_str.strip():
            return
        end_date_str = end_date_str.strip()
        if not re.match(r'^\d{4}-\d{2}-\d{2}$', end_date_str):
            messagebox.showwarning("형식 오류", "종료일 형식은 YYYY-MM-DD이어야 합니다.")
            return
        if end_date_str <= start_date_str:
            messagebox.showwarning("날짜 오류", "종료일은 시작일보다 이후여야 합니다.")
            return

        # 집계 전 출석체크/수강생후기/보충강의/라이브후기 값 초기화
        # 새로운 집계 결과로 덮어쓰기 위해 기존 집계 관련 컬럼을 0으로 리셋
        for row in self.original_data:
            # 행 길이가 고정 컬럼 수보다 짧으면 타입에 맞는 기본값으로 채움
            while len(row) < len(FIXED_COLUMNS):
                col_name = FIXED_COLUMNS[len(row)]
                st = FIXED_SORT_TYPES.get(col_name)
                if st == SORT_TYPE_NUMBER:
                    row.append('0')
                elif st == SORT_TYPE_COMPOSITE:
                    row.append('0(0/0)')
                else:
                    row.append('')
            row[COL_IDX['출석체크']] = '0'
            row[COL_IDX['수강생 후기(주차별/라스트원)']] = '0(0/0)'
            row[COL_IDX['보충강의']] = '0'
            row[COL_IDX['라이브 후기']] = ''
        self.data = [list(row) for row in self.original_data]
        self._refresh_table()
        self._auto_save()

        self.status_label.configure(text="게시판 집계 준비 중... (기존 집계값 초기화됨)")
        self.root.update()
        output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'cafe_members.json')

        # 별도 스레드에서 게시판 집계 스크래핑 실행
        def run():
            def on_status(msg):
                self.root.after(0, lambda: self.status_label.configure(text=msg))
            scraper = CafeScraper(config['cafe_name'], on_status=on_status)
            members = scraper.run_full(output_path, mode='boards',
                                       start_date=start_date_str, end_date=end_date_str,
                                       board_config=config)
            self.root.after(0, lambda: self._on_scrape_complete(members, output_path))

        threading.Thread(target=run, daemon=True).start()

    def _on_scrape_complete(self, members, output_path):
        if not members:
            messagebox.showinfo("결과", "수집된 회원이 없습니다.\n먼저 '회원 조회'를 실행하세요.")
            return
        if os.path.exists(output_path):
            self._load_cafe_json_from_path(output_path)
            try:
                with open(output_path, 'r', encoding='utf-8') as f:
                    payload = json.load(f)
                parts = [f"총 {len(members)}명"]
                for key, label in [('attendance_data', '출석체크'), ('weekly_review_data', '주차별후기'),
                                   ('lastone_review_data', '라스트원후기'), ('supplement_data', '보충강의'),
                                   ('live_review_data', '라이브후기')]:
                    d = payload.get(key, {})
                    if d:
                        parts.append(f"{label} {sum(d.values())}건")
                messagebox.showinfo("완료", ' / '.join(parts))
            except Exception:
                messagebox.showinfo("완료", f"총 {len(members)}명 처리됨")

    def _load_cafe_json_from_path(self, path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                payload = json.load(f)
            self._merge_cafe_json(payload)
        except Exception as e:
            messagebox.showerror("오류", f"JSON을 불러올 수 없습니다:\n{e}")

    # ──────────────────────────────────────────────
    # 카페 JSON 불러오기 / 병합
    # ──────────────────────────────────────────────

    def _load_cafe_json(self):
        path = filedialog.askopenfilename(
            title="카페 회원 JSON 선택",
            filetypes=[("JSON 파일", "*.json"), ("모든 파일", "*.*")])
        if path:
            self._load_cafe_json_from_path(path)

    def _merge_cafe_json(self, payload):
        """JSON payload에서 회원 데이터를 기존 테이블에 병합"""
        members = payload.get('members', [])
        if not members:
            messagebox.showinfo("알림", "회원 데이터가 없습니다.")
            return

        attendance_data = payload.get('attendance_data', {})
        weekly_review_data = payload.get('weekly_review_data', {})
        lastone_review_data = payload.get('lastone_review_data', {})
        supplement_data = payload.get('supplement_data', {})
        live_review_data = payload.get('live_review_data', {})

        id_index = {}
        for i, row in enumerate(self.original_data):
            nid = row[COL_IDX['네이버ID']].strip()
            if nid:
                id_index[nid] = i

        new_count = 0
        update_count = 0

        for m in members:
            nid = str(m.get('naver_id', '')).strip()
            nick = str(m.get('nickname', '')).strip()
            total_posts = int(m.get('post_count', 0))

            if not nid:
                continue

            if nid and nick:
                nick = re.sub(r'\s*\(\s*' + re.escape(nid) + r'\s*\)\s*', '', nick).strip()
                if not nick:
                    nick = nid

            att_count = attendance_data.get(nid, 0)
            adjusted_posts = str(max(total_posts - att_count, 0))
            att_str = str(att_count)

            weekly = weekly_review_data.get(nid, 0)
            lastone = lastone_review_data.get(nid, 0)
            review_str = f"{weekly + lastone}({weekly}/{lastone})"
            supplement_str = str(supplement_data.get(nid, 0))
            live_str = '◇' if live_review_data.get(nid, 0) > 0 else ''

            if nid in id_index:
                idx = id_index[nid]
                changed = False
                old = self.original_data[idx]
                # 행 길이 맞추기
                while len(old) < len(FIXED_COLUMNS):
                    old.append('')

                if old[COL_IDX['카페닉네임']] != nick:
                    old[COL_IDX['카페닉네임']] = nick
                    changed = True
                if old[COL_IDX['게시글수']] != adjusted_posts:
                    old[COL_IDX['게시글수']] = adjusted_posts
                    changed = True
                if att_str != '0' and old[COL_IDX['출석체크']] != att_str:
                    old[COL_IDX['출석체크']] = att_str
                    changed = True
                if review_str != '0(0/0)' and old[COL_IDX['수강생 후기(주차별/라스트원)']] != review_str:
                    old[COL_IDX['수강생 후기(주차별/라스트원)']] = review_str
                    changed = True
                if supplement_str != '0' and old[COL_IDX['보충강의']] != supplement_str:
                    old[COL_IDX['보충강의']] = supplement_str
                    changed = True
                if live_str and old[COL_IDX['라이브 후기']] != live_str:
                    old[COL_IDX['라이브 후기']] = live_str
                    changed = True

                if changed:
                    self.row_flags[nid] = {'new': False, 'updated': True}
                    update_count += 1
            else:
                new_row = [nid, nick, '', '', '', adjusted_posts, att_str, review_str, supplement_str, live_str]
                self.original_data.append(new_row)
                self.row_flags[nid] = {'new': True, 'updated': False}
                id_index[nid] = len(self.original_data) - 1
                new_count += 1

        self.data = [list(row) for row in self.original_data]
        self._reset_sort()
        self._update_count()
        self._auto_save()
        extras = []
        if attendance_data:
            extras.append(f"출석체크 {sum(attendance_data.values())}건")
        if weekly_review_data or lastone_review_data:
            extras.append(f"후기 {sum(weekly_review_data.values())+sum(lastone_review_data.values())}건")
        if supplement_data:
            extras.append(f"보충강의 {sum(supplement_data.values())}건")
        extra_msg = f", {', '.join(extras)} 반영" if extras else ""
        self.status_label.configure(
            text=f"카페 데이터 로드 완료: 신규 {new_count}명, 갱신 {update_count}명{extra_msg}")

    # ──────────────────────────────────────────────
    # CSV 속성 매핑 (동행천만, 밥상모임, 본명)
    # ──────────────────────────────────────────────

    def _infer_headers(self, raw_rows):
        """
        헤더 없는 CSV에서 데이터 패턴으로 열 이름 추론.
        - 한글만 평균 2~4글자 = 본명/이름
        - 영소문자+숫자 4~30자 = 네이버ID
        - 한글 포함 평균 4글자+ 또는 "id(닉네임)" 형식 = 카페닉네임
        - 전화번호 패턴 = 무시
        """
        if not raw_rows or len(raw_rows) < 2:
            return None

        num_cols = max(len(r) for r in raw_rows)
        if num_cols < 2:
            return None

        # 각 열의 샘플 분석 (최대 30행, 빈 행 제외)
        sample_rows = [r for r in raw_rows[:30] if any(c.strip() for c in r)]
        headers = [f'col_{i}' for i in range(num_cols)]
        assigned = set()  # 이미 할당된 열 이름 추적 (중복 방지)

        for col_idx in range(num_cols):
            # 해당 열의 비어있지 않은 값만 추출
            values = [r[col_idx].strip() for r in sample_rows if col_idx < len(r) and r[col_idx].strip()]
            if not values:
                continue

            # 전화번호 패턴 (010-xxxx-xxxx) → 유의미하지 않으므로 건너뛰기
            phone_count = sum(1 for v in values if re.match(r'^01\d[\-\s]?\d{3,4}[\-\s]?\d{4}$', v))
            if phone_count / len(values) > 0.5:
                headers[col_idx] = '전화번호'
                continue

            # 순수 숫자 열 (일련번호 등) → 건너뛰기
            if all(v.isdigit() for v in values):
                headers[col_idx] = '번호'
                continue

            # 한글만으로 구성, 평균 2~4글자 → 본명으로 추론
            korean_only = [v for v in values if re.match(r'^[가-힣]+$', v)]
            if korean_only and '본명' not in assigned:
                avg_len = sum(len(v) for v in korean_only) / len(korean_only)
                if 2 <= avg_len <= 4 and len(korean_only) / len(values) > 0.5:
                    headers[col_idx] = '본명'
                    assigned.add('본명')
                    continue

            # 영문+숫자 조합 → 네이버ID로 추론
            id_like = [v for v in values if re.match(r'^[a-zA-Z0-9_.\-@]+$', v)]
            if id_like and len(id_like) / len(values) > 0.3 and '네이버ID' not in assigned:
                headers[col_idx] = '네이버ID'
                assigned.add('네이버ID')
                continue

            # 한글 포함 + 평균 길이 3자 이상 또는 괄호 형식 → 카페닉네임으로 추론
            has_korean = [v for v in values if re.search(r'[가-힣]', v)]
            has_paren = [v for v in values if re.search(r'[\(\（]', v)]
            if (has_korean or has_paren) and '카페닉네임' not in assigned:
                avg_len = sum(len(v) for v in values) / len(values)
                if avg_len >= 3 or has_paren:
                    headers[col_idx] = '카페닉네임'
                    assigned.add('카페닉네임')
                    continue

        # 최소한 하나라도 유의미한 열을 찾았으면 반환
        if assigned:
            return headers
        return None

    def _read_csv_rows(self, csv_path):
        """CSV/XLSX 파일을 읽어 dict 리스트로 반환 / Read CSV or XLSX file and return list of row dicts"""
        # 확장자에 따라 Excel 파일이면 별도 처리
        _, ext = os.path.splitext(csv_path)
        if ext.lower() in ('.xlsx', '.xls'):
            return self._read_xlsx_rows(csv_path)
        # 인코딩 자동 감지: utf-8-sig 실패 시 euc-kr로 재시도
        try:
            with open(csv_path, 'r', encoding='utf-8-sig') as f:
                raw_rows = list(csv.reader(f))
        except UnicodeDecodeError:
            with open(csv_path, 'r', encoding='euc-kr') as f:
                raw_rows = list(csv.reader(f))
        if not raw_rows:
            return []
        # col_config에서 헤더 행 식별용 키워드 수집
        header_keywords = set()
        for cfg_key in ('id_keywords', 'nick_keywords', 'name_keywords'):
            for kw in self.col_config.get(cfg_key, '').split(','):
                kw = kw.strip()
                if kw:
                    header_keywords.add(kw)
        # 키워드가 포함된 행을 헤더 행으로 인식
        headers = None
        header_row_idx = -1
        for i, row in enumerate(raw_rows):
            stripped = [c.strip() for c in row]
            if any(s in header_keywords for s in stripped):
                headers = stripped
                header_row_idx = i
                break
        if not headers:
            # 헤더가 없는 CSV → 데이터 패턴으로 열 이름 자동 추론
            headers = self._infer_headers(raw_rows)
            if headers:
                header_row_idx = -1  # 헤더 행이 없으므로 0번 행부터 데이터
            else:
                # 추론 실패 시 DictReader 폴백
                try:
                    with open(csv_path, 'r', encoding='utf-8-sig') as f:
                        return list(csv.DictReader(f))
                except Exception:
                    return []
        # 헤더 이후 행들을 dict로 변환
        result = []
        for i in range(header_row_idx + 1, len(raw_rows)):
            row = raw_rows[i]
            # 빈 행 건너뛰기
            if not any(c.strip() for c in row):
                continue
            # 중간에 또 헤더 행이 나오면 건너뛰기 (반복 헤더 패턴 대응)
            stripped = [c.strip() for c in row]
            if any(s in header_keywords for s in stripped):
                continue
            row_dict = {}
            for j, val in enumerate(row):
                if j < len(headers) and headers[j]:
                    row_dict[headers[j]] = val.strip()
            # 키워드 열에 값이 하나라도 있는 행만 결과에 포함
            if any(row_dict.get(h, '') for h in headers if h in header_keywords):
                result.append(row_dict)
        return result

    def _read_xlsx_rows(self, xlsx_path):
        """Excel(.xlsx) 파일을 읽어 dict 리스트로 반환 / Read XLSX file and return list of row dicts"""
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("모듈 없음", "openpyxl이 필요합니다.")
            return []
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb[wb.sheetnames[0]]  # 첫 번째 시트만 사용
        # col_config에서 헤더 식별용 키워드 수집
        header_keywords = set()
        for cfg_key in ('id_keywords', 'nick_keywords', 'name_keywords'):
            for kw in self.col_config.get(cfg_key, '').split(','):
                kw = kw.strip()
                if kw:
                    header_keywords.add(kw)
        # 처음 50행 내에서 키워드가 포함된 헤더 행 탐색
        headers = None
        header_row = 0
        for row_idx in range(1, min(ws.max_row + 1, 50)):
            row_vals = [str(c.value).strip() if c.value else '' for c in ws[row_idx]]
            if any(v in header_keywords for v in row_vals):
                headers = row_vals
                header_row = row_idx
                break
        if not headers:
            return []
        # 헤더 이후 행들을 dict로 변환
        result = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_vals = [str(c.value).strip() if c.value else '' for c in ws[row_idx]]
            if not any(row_vals):
                continue
            # 반복 헤더 행 건너뛰기
            if any(v in header_keywords for v in row_vals):
                continue
            row_dict = {}
            for j, val in enumerate(row_vals):
                if j < len(headers) and headers[j]:
                    row_dict[headers[j]] = val
            if any(row_dict.get(h, '') for h in headers if h in header_keywords):
                result.append(row_dict)
        return result

    # ──────────────────────────────────────────────
    # 속성 매핑 열 이름 설정
    # ──────────────────────────────────────────────

    def _load_col_config(self):
        default = {
            "id_keywords": "네이버ID,네이버id,naver_id,ID,id,아이디,Id,iD",
            "nick_keywords": "카페닉네임,닉네임,별명,nickname",
            "name_keywords": "본명,이름,성명,성함,실명,활동명,name",
        }
        if os.path.exists(self.col_config_file):
            try:
                with open(self.col_config_file, 'r', encoding='utf-8') as f:
                    saved = json.load(f)
                for k in default:
                    if k not in saved:
                        saved[k] = default[k]
                return saved
            except Exception:
                pass
        return default

    def _save_col_config(self, config):
        with open(self.col_config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        self.col_config = config

    def _open_col_config(self):
        """속성 매핑 열 이름 설정 다이얼로그"""
        dialog = tk.Toplevel(self.root)
        dialog.title("속성 매핑 열 이름 설정")
        dialog.configure(bg='#1a1a2e')
        dialog.geometry("550x220")
        dialog.transient(self.root)
        dialog.grab_set()

        entries = {}
        fields = [
            ("네이버ID 열 키워드", "id_keywords"),
            ("카페닉네임 열 키워드", "nick_keywords"),
            ("본명 열 키워드", "name_keywords"),
        ]
        for i, (label, key) in enumerate(fields):
            tk.Label(dialog, text=label, font=('맑은 고딕', 10),
                     bg='#1a1a2e', fg='white').grid(row=i, column=0, padx=10, pady=8, sticky='e')
            entry = tk.Entry(dialog, font=('맑은 고딕', 10), bg='#16213e', fg='white',
                             insertbackground='white', width=45)
            entry.insert(0, self.col_config.get(key, ''))
            entry.grid(row=i, column=1, padx=10, pady=8, sticky='ew')
            entries[key] = entry
        dialog.grid_columnconfigure(1, weight=1)

        def save():
            new_config = {k: entries[k].get().strip() for k in entries}
            self._save_col_config(new_config)
            self.status_label.configure(text="열 이름 설정 저장됨")
            dialog.destroy()

        tk.Button(dialog, text="저장", command=save,
                  font=('맑은 고딕', 10, 'bold'), bg='#1a5276', fg='white',
                  relief='flat', padx=20, pady=4).grid(
            row=len(fields), column=0, columnspan=2, pady=12)

    def _find_csv_key(self, csv_keys, config_key, csv_rows=None):
        """
        CSV 헤더에서 열 이름 찾기.
        1단계: col_config에 설정된 키워드로 정확히 매칭
        2단계: 대소문자 무시 매칭
        3단계: 자동 패턴 탐색
        4단계: 데이터 기반 자동 추론 (폴백)
        """
        keywords = [kw.strip() for kw in self.col_config.get(config_key, '').split(',') if kw.strip()]

        # 1단계: 정확히 매칭
        for k in csv_keys:
            if k.strip() in keywords:
                return k

        # 2단계: 대소문자 무시 매칭
        keywords_lower = [kw.lower() for kw in keywords]
        for k in csv_keys:
            if k.strip().lower() in keywords_lower:
                return k

        # 3단계: 자동 패턴 탐색
        auto_patterns = {
            'id_keywords': lambda k: k.strip().lower() in ('id', '아이디') or 'id' in k.strip().lower(),
            'nick_keywords': lambda k: any(w in k.strip() for w in ('닉네임', '별명', 'nickname')),
            'name_keywords': lambda k: any(w in k.strip() for w in ('본명', '이름', '성명', '성함', '실명', '활동명', 'name')),
        }
        if config_key in auto_patterns:
            for k in csv_keys:
                if auto_patterns[config_key](k):
                    return k

        # 4단계: 데이터 기반 자동 추론 (csv_rows가 있을 때만)
        if csv_rows and len(csv_rows) >= 3:
            for k in csv_keys:
                values = [r.get(k, '').strip() for r in csv_rows[:30] if r.get(k, '').strip()]
                if not values:
                    continue

                if config_key == 'name_keywords':
                    # 한글만, 평균 약 3글자
                    korean_only = [v for v in values if re.match(r'^[가-힣]+$', v)]
                    if korean_only:
                        avg_len = sum(len(v) for v in korean_only) / len(korean_only)
                        if 2 <= avg_len <= 4 and len(korean_only) / len(values) > 0.5:
                            return k

                elif config_key == 'id_keywords':
                    # 영문소문자+숫자+특수문자 조합, 8-16자
                    id_like = [v for v in values if re.match(r'^[a-z0-9_.\-@]{4,30}$', v)]
                    if id_like and len(id_like) / len(values) > 0.5:
                        return k

                elif config_key == 'nick_keywords':
                    # 한글 포함, 평균 4글자 이상
                    korean_vals = [v for v in values if re.search(r'[가-힣]', v)]
                    if korean_vals:
                        avg_len = sum(len(v) for v in korean_vals) / len(korean_vals)
                        if avg_len >= 4 and len(korean_vals) / len(values) > 0.3:
                            return k

        return None

    @staticmethod
    def _split_nick_id(value):
        """
        닉네임과 ID 분리 (양방향 지원):
        "준준성수 (tntnkj)" → nick='준준성수', id='tntnkj'
        "spoon0 (트레이서)" → nick='트레이서', id='spoon0'
        "ryujs37767(다닝)"  → nick='다닝', id='ryujs37767'
        "eric"              → nick='eric', id=''
        """
        # 괄호 안 내용 추출
        match = re.match(r'^(.+?)\s*[\(\（](.+?)[\)\）]\s*$', value)
        if match:
            outside = match.group(1).strip()
            inside = match.group(2).strip()
            # 바깥이 영문+숫자(ID) / 안이 한글(닉네임)
            if re.match(r'^[a-zA-Z0-9_.\-@]+$', outside) and re.search(r'[가-힣]', inside):
                return inside, outside  # nick=안, id=바깥
            # 바깥이 한글(닉네임) / 안이 영문+숫자(ID)
            if re.match(r'^[a-zA-Z0-9_.\-@]+$', inside):
                return outside, inside  # nick=바깥, id=안
            # 둘 다 한글이면 바깥=닉네임, 안=추가정보
            return outside, inside
        return value, ''

    def _match_csv_to_rows(self, csv_path):
        csv_rows = self._read_csv_rows(csv_path)
        if not csv_rows:
            return {}

        sample_keys = list(csv_rows[0].keys())
        id_key = self._find_csv_key(sample_keys, 'id_keywords', csv_rows)
        nick_key = self._find_csv_key(sample_keys, 'nick_keywords', csv_rows)
        name_key = self._find_csv_key(sample_keys, 'name_keywords', csv_rows)

        # 기존 데이터 룩업 생성
        id_to_idx = {}
        nick_to_idx = {}           # 정확 매칭
        nick_lower_to_idx = {}     # 대소문자 무시
        nick_lower_count = {}      # 동음이인 카운트 (대소문자 무시)
        nick_nospace_to_idx = {}   # 띄어쓰기+대소문자 무시
        nick_nospace_count = {}    # 동음이인 카운트 (띄어쓰기 무시)
        name_to_idx = {}
        for i, row in enumerate(self.original_data):
            nid = row[COL_IDX['네이버ID']].strip()
            nick = row[COL_IDX['카페닉네임']].strip()
            name = row[COL_IDX['본명']].strip()
            if nid:
                id_to_idx[nid] = i
            if nick:
                nick_to_idx.setdefault(nick, i)
                nl = nick.lower()
                nick_lower_to_idx.setdefault(nl, i)
                nick_lower_count[nl] = nick_lower_count.get(nl, 0) + 1
                # 띄어쓰기 제거 키
                nns = nick.lower().replace(' ', '')
                nick_nospace_to_idx.setdefault(nns, i)
                nick_nospace_count[nns] = nick_nospace_count.get(nns, 0) + 1
            if name:
                name_to_idx.setdefault(name, i)

        result = {}
        for csv_row in csv_rows:
            matched_idx = None

            # "닉네임 (id)" 형식인 값이 있으면 분리해서 시도
            combined_nick = ''
            combined_id = ''

            # 우선순위 1: 네이버ID 열
            if id_key:
                val = csv_row.get(id_key, '').strip()
                if val in id_to_idx:
                    matched_idx = id_to_idx[val]

            # 우선순위 2: 카페닉네임 열
            if matched_idx is None and nick_key:
                val = csv_row.get(nick_key, '').strip()
                # "준준성수 (tntnkj)" 형태 분리
                nick_part, id_part = self._split_nick_id(val)

                # 닉네임으로 정확 매칭
                if nick_part in nick_to_idx:
                    matched_idx = nick_to_idx[nick_part]
                # 대소문자 무시 (동음이인 없을 때만)
                elif nick_part and nick_part.lower() in nick_lower_to_idx:
                    if nick_lower_count.get(nick_part.lower(), 0) == 1:
                        matched_idx = nick_lower_to_idx[nick_part.lower()]
                # 띄어쓰기+대소문자 무시 (동음이인 없을 때만)
                if matched_idx is None and nick_part:
                    nns = nick_part.lower().replace(' ', '')
                    if nns in nick_nospace_to_idx and nick_nospace_count.get(nns, 0) == 1:
                        matched_idx = nick_nospace_to_idx[nns]
                # 분리된 ID로 시도
                if matched_idx is None and id_part and id_part in id_to_idx:
                    matched_idx = id_to_idx[id_part]
                # 원본 값(분리 전)으로도 시도
                if matched_idx is None and val != nick_part:
                    if val in nick_to_idx:
                        matched_idx = nick_to_idx[val]

            # 우선순위 3: 본명
            if matched_idx is None and name_key:
                val = csv_row.get(name_key, '').strip()
                if val in name_to_idx:
                    matched_idx = name_to_idx[val]

            # 우선순위 4: 모든 열에서 "닉네임 (id)" 형식 찾기
            if matched_idx is None:
                for key, val in csv_row.items():
                    val = val.strip()
                    if not val:
                        continue
                    nick_part, id_part = self._split_nick_id(val)
                    if id_part and id_part in id_to_idx:
                        matched_idx = id_to_idx[id_part]
                        break
                    if nick_part and nick_part in nick_to_idx:
                        matched_idx = nick_to_idx[nick_part]
                        break

            if matched_idx is not None:
                result[matched_idx] = dict(csv_row)

        return result

    def _load_attribute_csv(self, title, target_col, col_keywords):
        path = filedialog.askopenfilename(
            title=f"{title} CSV 선택",
            filetypes=[("CSV/Excel 파일", "*.csv *.xlsx"), ("CSV 파일", "*.csv"),
                       ("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")])
        if not path:
            return
        matches = self._match_csv_to_rows(path)
        filled = 0
        for idx, csv_row in matches.items():
            value = ''
            for kw in col_keywords:
                for key in csv_row:
                    if kw in key.strip():
                        value = csv_row[key].strip()
                        break
                if value:
                    break
            if value:
                self.original_data[idx][COL_IDX[target_col]] = value
            else:
                self.original_data[idx][COL_IDX[target_col]] = 'O'
            filled += 1
        self.data = [list(row) for row in self.original_data]
        self._refresh_table()
        self._auto_save()
        total = sum(1 for row in self.original_data
                    if row[COL_IDX[target_col]].strip() not in ('', 'X'))
        self.status_label.configure(text=f"{title} 매칭: {filled}명 적용 (전체 {total}명 표시)")

    def _load_dongheng_csv(self):
        """동행천만 CSV를 불러와 '소속'에 '기'가 포함된 회원에 별표 표시 / Load Donghengcheonman CSV and mark members"""
        path = filedialog.askopenfilename(
            title="동행천만 멤버 CSV 선택",
            filetypes=[("CSV/Excel 파일", "*.csv *.xlsx"), ("CSV 파일", "*.csv"),
                       ("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")])
        if not path:
            return
        matches = self._match_csv_to_rows(path)
        filled = 0
        for idx, csv_row in matches.items():
            # '소속' 열에서 '기'(기수) 포함 여부로 동행천만 멤버 판별
            sosok_val = ''
            for key in csv_row:
                if '소속' in key.strip():
                    sosok_val = csv_row[key].strip()
                    break
            if '기' in sosok_val:
                self.original_data[idx][COL_IDX['동행천만']] = '★'
                filled += 1
        self.data = [list(row) for row in self.original_data]
        self._refresh_table()
        self._auto_save()
        total = sum(1 for row in self.original_data
                    if row[COL_IDX['동행천만']].strip() not in ('', 'X'))
        self.status_label.configure(text=f"동행천만 매칭: {filled}명 ★ 표시 (전체 {total}명)")

    def _load_babsang_csv(self):
        """밥상모임 CSV를 불러와 해당 회원에 표시 / Load Babsang CSV and mark matching members"""
        self._load_attribute_csv("밥상모임", '밥상모임', ['밥상모임', '밥상', '모임'])

    def _load_realname_csv(self):
        """본명 매핑 CSV를 불러와 회원별 본명 입력 / Load real-name CSV and fill in the name column"""
        path = filedialog.askopenfilename(
            title="본명 매핑 CSV 선택",
            filetypes=[("CSV/Excel 파일", "*.csv *.xlsx"), ("CSV 파일", "*.csv"),
                       ("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")])
        if not path:
            return
        matches = self._match_csv_to_rows(path)
        filled = 0

        # 본명 열 찾기 (설정 키워드 기반)
        csv_rows = self._read_csv_rows(path)
        name_col_key = None
        if csv_rows:
            name_col_key = self._find_csv_key(list(csv_rows[0].keys()), 'name_keywords')

        for idx, csv_row in matches.items():
            name_val = ''
            # 설정된 키로 먼저 시도
            if name_col_key and name_col_key in csv_row:
                name_val = csv_row[name_col_key].strip()
            # 폴백: 모든 본명 키워드로 시도
            if not name_val:
                name_kws = [kw.strip() for kw in self.col_config.get('name_keywords', '').split(',')]
                for key in name_kws:
                    if key in csv_row and csv_row[key].strip():
                        name_val = csv_row[key].strip()
                        break
            if name_val:
                # 괄호와 괄호 안 내용 제거: "송호상(토)" -> "송호상" (부가 정보 제거)
                name_val = re.sub(r'\s*[\(\（].*?[\)\）]\s*', '', name_val).strip()
                self.original_data[idx][COL_IDX['본명']] = name_val
                filled += 1
        self.data = [list(row) for row in self.original_data]
        self._refresh_table()
        self._auto_save()
        self.status_label.configure(text=f"본명 매칭: {len(matches)}명 중 {filled}명 본명 입력됨")

    # ──────────────────────────────────────────────
    # 행 추가 / 삭제
    # ──────────────────────────────────────────────

    def _add_row(self):
        """회원 추가 다이얼로그를 표시하고 새 행 삽입 / Show add-member dialog and insert a new row"""
        dialog = tk.Toplevel(self.root)
        dialog.title("회원 추가")
        dialog.configure(bg='#1a1a2e')
        dialog.geometry("450x400")
        dialog.transient(self.root)
        dialog.grab_set()
        entries = {}
        for i, col in enumerate(self.columns):
            tk.Label(dialog, text=col, font=('맑은 고딕', 10),
                     bg='#1a1a2e', fg='white').grid(row=i, column=0, padx=10, pady=4, sticky='e')
            if col in ('동행천만', '밥상모임'):
                var = tk.StringVar(value='X')
                frame = tk.Frame(dialog, bg='#1a1a2e')
                frame.grid(row=i, column=1, padx=10, pady=4, sticky='w')
                tk.Radiobutton(frame, text='O', variable=var, value='O',
                               bg='#1a1a2e', fg='white', selectcolor='#0f3460',
                               font=('맑은 고딕', 10)).pack(side='left', padx=(0, 10))
                tk.Radiobutton(frame, text='X', variable=var, value='X',
                               bg='#1a1a2e', fg='white', selectcolor='#0f3460',
                               font=('맑은 고딕', 10)).pack(side='left')
                entries[col] = var
            else:
                entry = tk.Entry(dialog, font=('맑은 고딕', 10), bg='#16213e', fg='white',
                                 insertbackground='white')
                entry.grid(row=i, column=1, padx=10, pady=4, sticky='ew')
                if col in ('게시글수', '출석체크', '보충강의'):
                    entry.insert(0, '0')
                elif col == '수강생 후기(주차별/라스트원)':
                    entry.insert(0, '0(0/0)')
                entries[col] = entry
        dialog.grid_columnconfigure(1, weight=1)

        def save():
            row = [entries[col].get() for col in self.columns]
            for num_col in ('게시글수', '출석체크', '보충강의'):
                try:
                    int(row[COL_IDX[num_col]].replace(',', ''))
                except ValueError:
                    messagebox.showwarning("입력 오류", f"{num_col}는 숫자여야 합니다.")
                    return
            self.original_data.append(row)
            self.data = [list(r) for r in self.original_data]
            self._refresh_table()
            self._update_count()
            self._auto_save()
            self.status_label.configure(text="회원 추가됨")
            dialog.destroy()

        tk.Button(dialog, text="추가", command=save,
                  font=('맑은 고딕', 10, 'bold'), bg='#e94560', fg='white',
                  relief='flat', padx=20, pady=4).grid(
            row=len(self.columns), column=0, columnspan=2, pady=10)

    def _delete_row(self):
        """선택된 행(들)을 삭제 / Delete selected row(s) from the table"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("선택 필요", "삭제할 행을 선택하세요.\n(Ctrl+클릭 또는 Shift+클릭으로 여러 행 선택 가능)")
            return
        count = len(selected)
        if count > 1:
            if not messagebox.askyesno("확인", f"선택된 {count}명을 삭제하시겠습니까?"):
                return
        to_delete = [list(self.tree.item(item, 'values')) for item in selected]
        for values in to_delete:
            for i, row in enumerate(self.original_data):
                if row == values or [str(v) for v in row] == values:
                    self.row_flags.pop(row[COL_IDX['네이버ID']], None)
                    self.original_data.pop(i)
                    break
        self.data = [list(row) for row in self.original_data]
        self._refresh_table()
        self._update_count()
        self._auto_save()
        self.status_label.configure(text=f"{count}명 삭제됨")

    def _delete_all_rows(self):
        """모든 행을 삭제 (확인 후 실행) / Delete all rows after user confirmation"""
        if not self.original_data:
            return
        count = len(self.original_data)
        if not messagebox.askyesno("모든 행 삭제",
                f"전체 {count}명의 데이터를 모두 삭제하시겠습니까?\n이 작업은 되돌릴 수 없습니다."):
            return
        self.original_data.clear()
        self.data.clear()
        self.row_flags.clear()
        self._refresh_table()
        self._update_count()
        self._auto_save()
        self.status_label.configure(text=f"전체 {count}명 삭제됨")

    # ──────────────────────────────────────────────
    # CSV 저장 / 자동 저장 / 로드
    # ──────────────────────────────────────────────

    def _save_csv(self):
        """현재 테이블 데이터를 CSV 파일로 내보내기 / Export current table data to a CSV file"""
        path = filedialog.asksaveasfilename(
            title="CSV로 저장", defaultextension=".csv",
            filetypes=[("CSV 파일", "*.csv")])
        if not path:
            return
        try:
            # 필터가 적용된 경우 화면에 보이는 행만 저장
            visible_rows = [row for row in self.data if self._passes_filter(row)]
            with open(path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(self.columns)
                writer.writerows(visible_rows)
            total = len(self.data)
            saved = len(visible_rows)
            if saved < total:
                self.status_label.configure(text=f"저장됨: {os.path.basename(path)} ({saved}/{total}명, 필터 적용)")
            else:
                self.status_label.configure(text=f"저장됨: {os.path.basename(path)} ({saved}명)")
        except Exception as e:
            messagebox.showerror("오류", f"저장할 수 없습니다:\n{e}")

    def _auto_save(self):
        """원본 데이터를 member_data.json에 자동 저장 / Auto-save original data to member_data.json"""
        payload = {'saved_at': datetime.datetime.now().isoformat(), 'members': self.original_data}
        try:
            with open(self.data_file, 'w', encoding='utf-8') as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _load_persisted_data(self):
        """앱 시작 시 member_data.json에서 저장된 데이터 복원 / Load previously saved data on app startup"""
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    payload = json.load(f)
                data = payload.get('members', [])
                if data:
                    # 행 길이가 고정 컬럼 수보다 짧으면 타입별 기본값으로 채움
                    for row in data:
                        while len(row) < len(FIXED_COLUMNS):
                            col_name = FIXED_COLUMNS[len(row)]
                            st = FIXED_SORT_TYPES.get(col_name)
                            if st == SORT_TYPE_NUMBER:
                                row.append('0')
                            elif st == SORT_TYPE_COMPOSITE:
                                row.append('0(0/0)')
                            else:
                                row.append('')
                    self._set_data(data)
                    self.status_label.configure(text=f"저장된 데이터 로드: {len(data)}명")
                    return
            except Exception:
                pass
        # 저장 파일이 없거나 로드 실패 시 빈 테이블로 시작
        self._set_data([])
        self.status_label.configure(text="데이터를 불러오세요 (카페 JSON 또는 행 추가)")


def main():
    root = tk.Tk()
    root.minsize(1100, 450)
    app = SortableTable(root)
    root.mainloop()


if __name__ == '__main__':
    main()
